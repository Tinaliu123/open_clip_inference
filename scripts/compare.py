# compare.py
import pandas as pd
import json
import argparse
import os
from openpyxl import load_workbook
import zipfile

def save_to_indiv(class_metrics,model_name,average_precision,average_recall,average_f1,top1_accuracy,indiv_filename) -> None:
    # 创建一个空的字典
   row1 = {}
   row2 = {}
   class_metrics_list1 = []
   class_metrics_list2 = []
# 将 model_name 添加到字典中
   row1["model_name"] = model_name

# 从 class_metrics 中提取类别和相应的 precision、recall、f1
   for metrics in class_metrics["class_metrics"]:
    label = metrics["Label"]
    row1[f"{label}_precision"] = metrics["Precision"]
    row1[f"{label}_recall"] = metrics["Recall"]
    row1[f"{label}_f1"] = metrics["F1"]

   # 将字典添加到列表
   class_metrics_list1.append(row1)
   
   for metrics in class_metrics["class_metrics"]:
    row = {"class": metrics["Label"]}
    row[f"{model_name}_precision"] = metrics["Precision"]
    row[f"{model_name}_recall"] = metrics["Recall"]
    row[f"{model_name}_f1"] = metrics["F1"]

    # 将当前行添加到列表中
    class_metrics_list2.append(row)

   row2["class"] = "Average"
   row2[f"{model_name}_precision"] = average_precision
   row2[f"{model_name}_recall"] = average_recall
   row2[f"{model_name}_f1"] = average_f1
   class_metrics_list2.append(row2)
  

   metrics1_df = pd.DataFrame(class_metrics_list1)
   metrics2_df = pd.DataFrame(class_metrics_list2)
 
   overall_metrics1_df = pd.DataFrame({
    "Average_Precision": [average_precision],
    "Average_Recall": [average_recall],
    "Average_F1": [average_f1],
    "Top1_Accuracy": [top1_accuracy]
   })



    # 合并 metrics_df 和 overall_metrics_df
   final1 = pd.concat([metrics1_df, overall_metrics1_df], axis=1)
   final2 = metrics2_df
  
   with pd.ExcelWriter(indiv_filename, engine='openpyxl') as writer:
    # 写入第一个表
    final1.to_excel(writer, sheet_name='Sheet1', index=False)
    # 写入第二个表
    final2.to_excel(writer, sheet_name='Sheet2', index=False)
   print(f"individual result saved to {indiv_filename}")



def save_to_overall(class_metrics,model_name,average_precision,average_recall,average_f1,top1_accuracy,overall_filename) -> None:

      # 创建一个空的字典
   row1 = {}
   row2 = {}
   class_metrics_list1 = []
   class_metrics_list2 = []
# 将 model_name 添加到字典中
   row1["model_name"] = model_name

# 从 class_metrics 中提取类别和相应的 precision、recall、f1
   for metrics in class_metrics["class_metrics"]:
    label = metrics["Label"]
    row1[f"{label}_precision"] = metrics["Precision"]
    row1[f"{label}_recall"] = metrics["Recall"]
    row1[f"{label}_f1"] = metrics["F1"]

   # 将字典添加到列表
   class_metrics_list1.append(row1)

   for metrics in class_metrics["class_metrics"]:
    row = {"class": metrics["Label"]}
    row[f"{model_name}_precision"] = metrics["Precision"]
    row[f"{model_name}_recall"] = metrics["Recall"]
    row[f"{model_name}_f1"] = metrics["F1"]

    # 将当前行添加到列表中
    class_metrics_list2.append(row)
   row2["class"] = "Average"
   row2[f"{model_name}_precision"] = average_precision
   row2[f"{model_name}_recall"] = average_recall
   row2[f"{model_name}_f1"] = average_f1
   class_metrics_list2.append(row2)


   metrics1_df = pd.DataFrame(class_metrics_list1)
   metrics2_df = pd.DataFrame(class_metrics_list2)

   overall_metrics1_df = pd.DataFrame({
    "Average_Precision": [average_precision],
    "Average_Recall": [average_recall],
    "Average_F1": [average_f1],
    "Top1_Accuracy": [top1_accuracy]
   })


   final1 = pd.concat([metrics1_df, overall_metrics1_df], axis=1)
   final2 = metrics2_df
 
   if zipfile.is_zipfile(overall_filename):
        # 如果文件存在，读入原有的内容
        book = load_workbook(overall_filename)
        #writer = pd.ExcelWriter(overall_filename, engine='openpyxl', mode='a', if_sheet_exists='replace', index=False)
        #writer.book = book

        #表单一
        existing_data1 = pd.read_excel(overall_filename, sheet_name='Sheet1')
        final1 = pd.concat([existing_data1, final1], axis=0, ignore_index=True)
        sheet1 = book['Sheet1']
        book.remove(sheet1)
      
        #表单二合并需要处理
        existing_data2 = pd.read_excel(overall_filename, sheet_name='Sheet2')
        final2 = pd.merge(existing_data2, final2.iloc[:, 1:], left_index=True, right_index=True, how='outer')
        sheet2 = book['Sheet2']
        book.remove(sheet2)

     
   # else:
   #      # 如果文件不存在，直接创建一个新文件
   #      writer = pd.ExcelWriter(overall_filename, engine='openpyxl')
 
   with pd.ExcelWriter(overall_filename, engine='openpyxl') as writer:
    final1.to_excel(writer, sheet_name='Sheet1', index=False)
    final2.to_excel(writer, sheet_name='Sheet2', index=False)

   # 冻结
   excel = load_workbook(overall_filename)
   sheet1 = excel['Sheet1']
   sheet2 = excel['Sheet2']
   sheet1.freeze_panes='B2'
   sheet2.freeze_panes='B2'
   excel.save(overall_filename)
   print(f"Overall result saved to {overall_filename}")




def main(class_metrics_path, average_f1_top1_accuracy_path,model_name,indiv_result,overall_result) -> None:
   with open(class_metrics_path, "r") as class_metrics_file:
        class_metrics = json.load(class_metrics_file)

   with open(average_f1_top1_accuracy_path, "r") as average_f1_top1_accuracy_file:
        result_dict = json.load(average_f1_top1_accuracy_file)
   average_precision = result_dict["average_precision"]
   average_recall = result_dict["average_recall"]
   average_f1 = result_dict["average_f1"]
   top1_accuracy = result_dict["top1_accuracy"]

   save_to_indiv(class_metrics,model_name,average_precision,average_recall,average_f1,top1_accuracy,indiv_result)
   save_to_overall(class_metrics,model_name,average_precision,average_recall,average_f1,top1_accuracy,overall_result)
   
    

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="compare script")
    parser.add_argument("--model_name", type=str, default="ViT-B-32", help="Model name")
    parser.add_argument("--model_path", type=str, default=None, help="Path to the model weights")
    parser.add_argument("--dataset_path", type=str, default=None, help="Path to the dataset")
    args = parser.parse_args()
      # 提取 model_path 和 dataset_path 的最后一部分
    if args.model_path is not None:
      model_name_with_extension = os.path.basename(args.model_path)
      model_name, extension = os.path.splitext(model_name_with_extension)
    else:
      model_name = "default_model"
      # model_name="open_clip_VIT-B-32"

    if args.dataset_path is not None:
      dataset_name = os.path.basename(args.dataset_path)
    else:
      dataset_name = "default_dataset"
      #dataset_name = "EuroSAT_RGB"
        # 获取当前脚本所在文件夹的路径
    current_script_path = os.path.dirname(os.path.abspath(__file__))
     # 获取上一级的路径
    parent_path = os.path.dirname(current_script_path)
    intermediate = os.path.join(parent_path,"intermediate")
    results = os.path.join(parent_path,"results")
    intermediate_path = os.path.join(intermediate, f"{model_name}_{dataset_name}")
    class_metrics_path = os.path.join(intermediate_path,"recall_precision_f1.json")
    average_f1_top1_accuracy_path = os.path.join(intermediate_path,"avgf1_top1acc.json")
    indiv_result_path= os.path.join(intermediate_path,"indiv_result.xlsx")
    overall_result_filename=os.path.join(results,f"{dataset_name}.xlsx")
    main(class_metrics_path, average_f1_top1_accuracy_path,model_name,indiv_result_path,overall_result_filename)
